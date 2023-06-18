# modules needed for the device
from zaber.serial import BinarySerial, BinaryDevice
import nidmm

# modules needed for the graph
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# modules needed for the Gui
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QFont, QKeySequence
from PyQt5.QtCore import Qt

# additional modules to meet our need
import math
import openpyxl as excel
import os
import numpy as np

# this file_path is used to check
# if the excel file exists
file_path = r'C:\Users\louis\Desktop\final\lab.xlsx'


# a class to connect with the main Gui interface
class NewWindow(QDialog):
    def __init__(self, number: int):
        super().__init__()

        # initialize the sub Gui
        self.setWindowTitle("Picture")
        self.setGeometry(660, 240, 1200, 600)

        # create a space for the figure drawn afterwards
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)

        # the formation of the layout: vertical permutation
        layout = QVBoxLayout(self)
        # in this Gui the widget will be put in this section of the layout
        layout.addWidget(self.canvas)
        layout.addWidget(self.toolbar)

        # set the port of the connected device
        with BinarySerial("COM7") as port:
            # devide degree of 360 into number pieces
            deg = int(4267 * 360 / number)

            # initialize the device
            device = BinaryDevice(port, 1)

            # the check if the device is homw
            reply = device.home()
            if self.check_command_succeeded(reply):
                print("Device homed.")
            else:
                print("Device home failed.")
                exit(1)

            # two list to store degree and measurements
            lst_degree, lst_measure = [], []
            # start recording
            for N in range(1, number + 1):
                # move_rel makes the device to rotate a degree deg
                reply = device.move_rel(deg)
                # record the current degree and measurement
                lst_degree.append(round(deg * N / 4267, 2))
                with nidmm.Session("Dev1") as session:
                    lst_measure.append(session.read())

        # plot the data detected
        self.plot(lst_degree, lst_measure)

        # write all the data in an excel file
        if os.path.exists(file_path):
            # load the existing file
            Wb = excel.load_workbook('lab.xlsx')
            # choose the sheet that stores out data
            Ws = Wb['Data']

            # below shows all the data
            # transformed to the primitive data types
            # and take all of the data
            Datas = []
            for row in Ws:
                temp = []
                for elem in row:
                    string = str(elem)
                    # this tells us the position of the data
                    pos = string[string.index('.') + 1: -1]
                    # take the value of it with the attribute
                    if (val := Ws[pos].value) != None:
                        temp.append(val)
                Datas.append(temp)
            Wb.close()

            # new information will be put in
            # the number of information will increase by 1
            Datas[0][0] += 1
            # the form we want to store
            content = [
                [],
                ['deg', *lst_degree],
                ['mea', *lst_measure],
                ['Max', lst_measure.index(tmp := max(lst_measure)), tmp],
                ['Min', lst_measure.index(tmp := min(lst_measure)), tmp]
            ]
            # list contanenation
            Datas += content

            # below rewrite the excel file
            newWb = excel.Workbook()
            newWs = newWb.active
            newWs.title = 'Data'
            for row in Datas:
                # write in the excel file
                newWs.append(row)
            newWb.save('lab.xlsx')
        else:
            # create an excel file
            newWb = excel.Workbook()
            # active helps us to name the new sheet
            newWs = newWb.active
            newWs.title = 'Data'
            """
                the form we want to put in excel:

                number_of_Datas

                deg | ...
                mea | ...
                Max | deg_to_mea_max | mea_max
                Min | deg_to_mea_min | mea_min
            """
            content = [
                [1],
                [],
                ['deg', *lst_degree],
                ['mea', *lst_measure],
                ['Max', lst_measure.index(tmp := max(lst_measure)), tmp],
                ['Min', lst_measure.index(tmp := min(lst_measure)), tmp]
            ]
            # put datas in excel
            for row in content:
                newWs.append(row)
            # save it
            newWb.save('lab.xlsx')

    # plot the point detected by the multimeter
    def plot(self, lst_x: list, lst_y: list):
        max_y, min_y = math.ceil(max(lst_y)), math.floor(min(lst_y))
        # create a layout containing a subplot
        # subplot(row, col, index)
        xy_graph = self.figure.add_subplot(1, 2, 1)

        # set the x-axis tick value to let data and user know
        # where the information of point is
        xy_graph.set_xticks(lst_x)
        xy_graph.set_xticklabels(lst_x)

        # set the y-axis tick value
        xy_graph_ylbls = []
        if max_y - min_y <= 4:
            # if the mplitude is much smaller
            # the difference of each tick will be smaller
            m, M = min_y * 100, max_y * 100

            for elem in range(m, M + 1, 25):
                xy_graph_ylbls.append(elem / 100)
            xy_graph.set_yticks(xy_graph_ylbls)
        else:
            # if the amplitude is much larger
            # the difference of each tick will be larger
            m, M = min_y, max_y
            for elem in range(m, M + 1):
                xy_graph_ylbls.append(elem)
            xy_graph.set_yticks(xy_graph_ylbls)
        xy_graph.set_yticklabels(xy_graph_ylbls)

        # draw the xy-graph and the attributes
        xy_graph.set_title("XY Chart: ° vs Mea")
        xy_graph.set_xlabel("degree (°)")
        xy_graph.set_ylabel("Intensity")

        # draw on the figure
        xy_graph.plot(lst_x, lst_y)

        # plot a radar chart
        radar_chart = self.figure.add_subplot(1, 2, 2, polar=True)

        # this is the rad list corresponding to the measurements
        lst_rad = [elem * np.pi / 180 for elem in lst_x]
        radar_chart.set_xticks(lst_rad)
        radar_chart.set_xticklabels(lst_x)

        # set the radius of the chart
        radar_chart.set_ylim(0, max_y)

        # the chart attribure
        radar_chart.set_title("Radar Chart")
        radar_chart.plot(lst_rad, lst_y)

        # renew the picture
        self.canvas.draw()

    def check_command_succeeded(self, reply):
        # 255 is the binary error response code
        if reply.command_number == 255:
            # warning messagebox
            QMessageBox.warning(
                self, "Warning", "Danger! Command rejected.Error code: " + str(reply.data))
            return False
        else:
            return True


class MyWindow(QMainWindow):
    # we use it to prevent user
    # misinput too many times
    # later it will shutdown the gui automatically
    __cnt = 0

    def __init__(self):
        super().__init__()

        # build window and the attributes
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.setWindowTitle("Malus Experiment Test")

        # select main layout
        layout = QVBoxLayout(central_widget)

        # put sub layout in main layout
        subLayout_1 = QHBoxLayout()
        layout.addLayout(subLayout_1)

        # sub lay out widgets
        lbl_N_Datas = QLabel()
        lbl_N_Datas.setText("Data N =")
        # set the font perference
        lbl_N_Datas.setFont(QFont('Consolas', 20))
        subLayout_1.addWidget(lbl_N_Datas)

        # a single line textbox with attributes initializing
        self.txt_N_Datas = QLineEdit()
        self.txt_N_Datas.setFixedWidth(70)
        self.txt_N_Datas.setFixedHeight(40)
        # the length of the input is limited within 3 chars
        self.txt_N_Datas.setMaxLength(3)
        self.txt_N_Datas.setFont(QFont('Consolas', 16))
        subLayout_1.addWidget(self.txt_N_Datas)

        # a button with attributes initializing
        btn_submit = QPushButton()
        btn_submit.setText("SUBMIT")
        btn_submit.setFont(QFont('Consolas', 20))
        btn_submit.setFixedHeight(40)
        btn_submit.setFixedWidth(140)
        subLayout_1.addWidget(btn_submit)

        # button clicked and the number of the datas
        # will be a key to trigger the sub window
        btn_submit.clicked.connect(self.btn_submit_clicked)

        # if user press "Enter", it also helps user
        # to trigger the sub window
        shortcut_btn_submit = QShortcut(QKeySequence(Qt.Key_Return), self)
        shortcut_btn_submit.activated.connect(self.btn_submit_clicked)

        N_test, deg_to_max_mea, deg_to_min_mea, max_mea, min_mea = 0, 0, 0, 0, 0
        # if the file has already existed
        # grab the datas above
        if os.path.exists(file_path):
            Wb = excel.load_workbook('lab.xlsx')
            Ws = Wb['Data']

            Datas = []
            for row in Ws:
                temp = []
                for elem in row:
                    string = str(elem)
                    pos = string[string.index('.') + 1: -1]
                    if (val := Ws[pos].value) != None:
                        temp.append(val)
                Datas.append(temp)
            Wb.close()

            N_test = Datas[0][0]
            for row in Datas:
                if len(row) > 0:
                    if row[0] == 'Max':
                        deg_to_max_mea += row[1]
                        max_mea += row[2]
                    elif row[0] == 'Min':
                        deg_to_min_mea += row[1]
                        min_mea += row[2]
                    else:
                        continue
                else:
                    continue

        # sublayout 2
        subLayout_2 = QHBoxLayout()
        layout.addLayout(subLayout_2)

        # sub lay out widgets
        self.lbl_avg_max = QLabel()
        # a hint to tell user the current state
        self.lbl_avg_max.setText(
            f"Current {N_test} average max: deg = {deg_to_max_mea}, val = {max_mea}")
        self.lbl_avg_max.setFont(QFont('Consolas', 12))
        subLayout_2.addWidget(self.lbl_avg_max)

        # sublayout 2
        subLayout_3 = QHBoxLayout()
        layout.addLayout(subLayout_3)

        # sub lay out widgets
        self.lbl_avg_min = QLabel()
        # a hint to tell user the current state
        self.lbl_avg_min.setText(
            f"Current {N_test} average min: deg = {deg_to_min_mea}, val = {min_mea}")
        self.lbl_avg_min.setFont(QFont('Consolas', 12))
        subLayout_3.addWidget(self.lbl_avg_min)

    # to update the current state of the datas
    # and renew the statements on the main window
    def updateData(self):
        # read data from the lab.xlsx
        Wb = excel.load_workbook('lab.xlsx')
        Ws = Wb['Data']

        # grab all the data
        # and take the useful information to
        # calculate the deg and mea
        Datas = []
        for row in Ws:
            temp = []
            for elem in row:
                string = str(elem)
                pos = string[string.index('.') + 1: -1]
                if (val := Ws[pos].value) != None:
                    temp.append(val)
            Datas.append(temp)
        Wb.close()

        # some methods to calculate the information
        # and renew the text to tell user
        N_test, max_mea, min_mea, deg_to_max_mea, deg_to_min_mea = Datas[0][0], 0, 0, 0, 0
        for row in Datas:
            if len(row) > 0:
                if row[0] == 'Max':
                    deg_to_max_mea += row[1]
                    max_mea += row[2]
                elif row[0] == 'Min':
                    deg_to_min_mea += row[1]
                    min_mea += row[2]
                else:
                    continue
            else:
                continue
        self.lbl_avg_max.setText(
            f"Current {N_test} average max: deg = {deg_to_max_mea / N_test}, val = {max_mea / N_test}")
        self.lbl_avg_min.setText(
            f"Current {N_test} average min: deg = {deg_to_min_mea / N_test}, val = {min_mea / N_test}")

    # trigger the sub window or messagebox with inputs
    def btn_submit_clicked(self):
        # take the data from the textbox
        number = self.txt_N_Datas.text()

        # if the data is not formed in digits
        # wrong input because not a positive integer
        if not number.isdigit():
            # jump out the messagebox with bold 'Wrong Input'
            mes = QMessageBox()
            tempFont = QFont('Consolas', 16)
            tempFont.setBold(True)
            mes.setWindowTitle("Warning")
            mes.setText("Wrong Input!")
            mes.setFont(tempFont)
            mes.setStyleSheet(
                "QMessageBox QLabel { color: red; } QMessageBox QPushButton { color: black }")
            mes.setStandardButtons(QMessageBox.Ok)
            mes.exec_()
            self.__cnt += 1
            if self.__cnt == 3:
                # the user must play this gui
                # up to 3 times, the Gui will shutdown
                self.__cnt += 1
                if self.__cnt != 3:
                    QMessageBox.warning(self, "Warning",
                                        "Please re-input the N, and 100 <= N <= 180!")
                    self.txt_N_Datas.clear()
                else:
                    QMessageBox.warning(self, "Fuck-you", "FUCK-YOU")
                    exit()

        # if the input is integer but is too large or too small
        # set the integer to default
        elif int(number) < 90 or int(number) > 360:
            # jump out the notice if the user
            # want to reinput or use the default value
            mes = QMessageBox()
            tempFont = QFont('Consolas', 16)
            mes.setText("We'll set N = 150 for simplicity!")
            mes.setWindowTitle("Notice")
            mes.setStyleSheet(
                "QMessageBox QLabel { color: red; } QMessageBox QPushButton { color: black }")
            mes.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            result = mes.exec_()
            if result == QMessageBox.Yes:
                # take the default value of number = 180
                # update the information afterwards
                self.hide()
                newWindow = NewWindow(180)
                newWindow.exec_()
                self.update()
                self.show()
            else:
                # the user must play this gui
                # up to 3 times, the Gui will shutdown
                self.__cnt += 1
                if self.__cnt != 3:
                    QMessageBox.warning(self, "Warning",
                                        "Please re-input the N, and 100 <= N <= 180!")
                    self.txt_N_Datas.clear()
                else:
                    # a nice greeting that you mistreat the program
                    QMessageBox.warning(self, "Fuck-you", "FUCK-YOU")
                    exit()
        else:
            # hide tha main window and show the sub window
            self.hide()
            N = int(self.txt_N_Datas.text())
            newWindow = NewWindow(N)
            newWindow.exec_()
            self.updateData()
            self.show()


if __name__ == '__main__':
    app = QApplication([])
    window = MyWindow()
    window.show()
    app.exec_()
